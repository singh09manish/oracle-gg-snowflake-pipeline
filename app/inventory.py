"""
Read and write the table inventory from/to Excel (.xlsx) or CSV.

Expected columns (case-insensitive):
  - schema            (required)  — Oracle source schema
  - table_name        (required)  — Oracle table name
  - group_id          (required)  — Extract group assignment
  - target_schema     (optional)  — Snowflake target schema (defaults to source schema)
  - target            (optional)  — Replication target: snowflake | rds | both (defaults to snowflake)
  - rds_target_schema (optional)  — RDS target schema (defaults to source schema)
  - enabled           (optional)  — Y/N toggle (defaults to Y if missing)
  - disabled_reason   (optional)  — Why the table was disabled
  - disabled_at       (optional)  — ISO timestamp of when it was disabled
  - excluded_columns  (optional)  — Comma-separated columns to exclude (PII/PCI)

Any extra columns are silently ignored.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from app.models import TableInfo

log = logging.getLogger(__name__)

# Column name aliases (lowercase) → canonical field
_SCHEMA_ALIASES = {"schema", "schema_name", "owner", "source_schema"}
_TABLE_ALIASES = {"table_name", "tablename", "table", "name", "object_name"}
_TARGET_ALIASES = {"target_schema", "target_schema_name", "tgt_schema", "snowflake_schema"}
_GROUP_ALIASES = {"group_id", "group", "extract_group"}
_ENABLED_ALIASES = {"enabled", "active", "status"}
_REASON_ALIASES = {"disabled_reason", "reason", "disable_reason"}
_DISABLED_AT_ALIASES = {"disabled_at", "disabled_date", "disabled_timestamp"}
_EXCLUDED_COLS_ALIASES = {"excluded_columns", "exclude_columns", "colsexcept", "pii_columns"}
_TARGET_TYPE_ALIASES = {"target", "target_type", "replication_target", "destination"}
_RDS_TARGET_SCHEMA_ALIASES = {"rds_target_schema", "rds_schema", "target_rds_schema", "oracle_target_schema"}


def read_inventory(path: str | Path) -> List[TableInfo]:
    """
    Read an Excel (.xlsx) or CSV file and return deduplicated TableInfo list.

    Raises ValueError if required columns are missing or file is empty.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Inventory file not found: {path}")

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        rows = _read_excel(path)
    elif ext in (".csv", ".tsv"):
        rows = _read_csv(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}  (use .xlsx or .csv)")

    tables = _parse_rows(rows, str(path))
    tables = _deduplicate(tables)
    _validate(tables, str(path))
    return tables


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def _read_excel(path: Path) -> List[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to read .xlsx files.\n"
            "  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active sheet in {path}")

    rows_iter = ws.iter_rows(values_only=True)
    header_raw = next(rows_iter, None)
    if header_raw is None:
        raise ValueError(f"Empty spreadsheet: {path}")

    header = [str(c).strip().lower() if c else "" for c in header_raw]
    data: List[dict] = []
    for row in rows_iter:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip blank rows
        data.append({header[i]: (str(row[i]).strip() if row[i] is not None else "")
                      for i in range(min(len(header), len(row)))})
    wb.close()
    return data


def _read_csv(path: Path) -> List[dict]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [
            {k.strip().lower(): (v.strip() if v else "") for k, v in row.items() if k}
            for row in reader
        ]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _resolve_col(row: dict, aliases: set) -> Optional[str]:
    """Find first matching alias in row keys, return value."""
    for key in row:
        if key in aliases:
            return row[key]
    return None


def _parse_rows(rows: List[dict], source: str) -> List[TableInfo]:
    if not rows:
        raise ValueError(f"No data rows found in {source}")

    # Validate that required columns exist
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())

    found_schema = all_keys & _SCHEMA_ALIASES
    found_table = all_keys & _TABLE_ALIASES
    found_group = all_keys & _GROUP_ALIASES
    if not found_schema:
        raise ValueError(
            f"Missing schema column in {source}.\n"
            f"  Expected one of: {_SCHEMA_ALIASES}\n"
            f"  Found columns: {sorted(all_keys)}"
        )
    if not found_table:
        raise ValueError(
            f"Missing table_name column in {source}.\n"
            f"  Expected one of: {_TABLE_ALIASES}\n"
            f"  Found columns: {sorted(all_keys)}"
        )
    if not found_group:
        raise ValueError(
            f"Missing group_id column in {source}.\n"
            f"  Expected one of: {_GROUP_ALIASES}\n"
            f"  Found columns: {sorted(all_keys)}"
        )

    has_enabled_col = bool(all_keys & _ENABLED_ALIASES)

    tables: List[TableInfo] = []
    for i, row in enumerate(rows, start=2):  # row 2 = first data row in Excel
        schema = _resolve_col(row, _SCHEMA_ALIASES) or ""
        table = _resolve_col(row, _TABLE_ALIASES) or ""
        target = _resolve_col(row, _TARGET_ALIASES) or ""
        group_raw = _resolve_col(row, _GROUP_ALIASES) or ""

        if not schema or not table:
            log.warning("Row %d: skipping — empty schema or table_name", i)
            continue

        if not group_raw:
            raise ValueError(f"Row {i}: missing group_id for {schema}.{table}")
        try:
            group_id = int(group_raw)
        except ValueError:
            raise ValueError(f"Row {i}: group_id must be an integer, got '{group_raw}'")

        # Enabled/disabled toggle (defaults to Y if column is absent)
        enabled_raw = _resolve_col(row, _ENABLED_ALIASES) if has_enabled_col else "Y"
        enabled = (enabled_raw or "Y").strip().upper() not in ("N", "NO", "FALSE", "0", "DISABLED")

        disabled_reason = _resolve_col(row, _REASON_ALIASES) or ""
        disabled_at = _resolve_col(row, _DISABLED_AT_ALIASES) or ""

        # COLSEXCEPT — comma-separated list of columns to exclude (PII/PCI)
        excluded_raw = _resolve_col(row, _EXCLUDED_COLS_ALIASES) or ""
        excluded_columns = [
            c.strip().upper() for c in excluded_raw.split(",") if c.strip()
        ]

        # Target type — snowflake | rds | both (defaults to snowflake)
        target_type = (_resolve_col(row, _TARGET_TYPE_ALIASES) or "snowflake").strip().lower()
        if target_type not in ("snowflake", "rds", "both"):
            log.warning("Row %d: invalid target '%s' for %s.%s — defaulting to 'snowflake'",
                        i, target_type, schema, table)
            target_type = "snowflake"

        # RDS target schema — defaults to source schema
        rds_target_schema = _resolve_col(row, _RDS_TARGET_SCHEMA_ALIASES) or ""

        tables.append(TableInfo(
            schema=schema, name=table, group_id=group_id, target_schema=target,
            enabled=enabled, disabled_reason=disabled_reason, disabled_at=disabled_at,
            excluded_columns=excluded_columns,
            target=target_type, rds_target_schema=rds_target_schema,
        ))
    return tables


def _deduplicate(tables: List[TableInfo]) -> List[TableInfo]:
    seen: dict[str, TableInfo] = {}
    dupes = 0
    for t in tables:
        if t.fqn in seen:
            dupes += 1
            continue
        seen[t.fqn] = t
    if dupes:
        log.warning("Removed %d duplicate table entries", dupes)
    return list(seen.values())


def _validate(tables: List[TableInfo], source: str) -> None:
    if not tables:
        raise ValueError(f"No valid tables found in {source}")

    enabled = [t for t in tables if t.enabled]
    disabled = [t for t in tables if not t.enabled]
    schemas = {t.schema for t in enabled}
    log.info(
        "Loaded %d tables across %d schemas from %s: %s",
        len(tables), len(schemas), source, ", ".join(sorted(schemas)),
    )

    # Target distribution
    sf_count = sum(1 for t in enabled if t.goes_to_snowflake)
    rds_count = sum(1 for t in enabled if t.goes_to_rds)
    both_count = sum(1 for t in enabled if t.target == "both")
    if rds_count > 0:
        log.info(
            "  Targets: %d → Snowflake, %d → RDS, %d → Both",
            sf_count - both_count, rds_count - both_count, both_count,
        )

    if disabled:
        log.info(
            "  %d tables DISABLED (will be excluded from .prm generation)",
            len(disabled),
        )
        for t in disabled:
            reason = f" — {t.disabled_reason}" if t.disabled_reason else ""
            log.info("    %s%s", t.fqn, reason)


# ---------------------------------------------------------------------------
# Write inventory back to Excel (for enable/disable operations)
# ---------------------------------------------------------------------------

def write_inventory(tables: List[TableInfo], path: str | Path) -> None:
    """
    Write the full table list back to Excel, preserving all fields
    including enabled/disabled_reason/disabled_at.

    This is used by the `disable` and `enable` CLI commands to update
    the inventory in place.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to write .xlsx files.\n  pip install openpyxl")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "table_inventory"

    # Header row
    headers = [
        "schema", "table_name", "target_schema", "group_id", "target",
        "rds_target_schema", "enabled", "disabled_reason", "disabled_at",
        "excluded_columns",
    ]
    ws.append(headers)

    # Data rows
    for t in sorted(tables, key=lambda x: (x.group_id, x.schema, x.name)):
        ws.append([
            t.schema,
            t.name,
            t.target_schema,
            t.group_id,
            t.target,
            t.rds_target_schema,
            "Y" if t.enabled else "N",
            t.disabled_reason,
            t.disabled_at,
            ", ".join(t.excluded_columns) if t.excluded_columns else "",
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    # Conditional formatting — highlight disabled rows in red
    from openpyxl.styles import PatternFill, Font
    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    red_font = Font(color="CC0000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        enabled_cell = row[6]  # 0-indexed: column G = "enabled"
        if str(enabled_cell.value).upper() == "N":
            for cell in row:
                cell.fill = red_fill
                cell.font = red_font

    wb.save(path)
    log.info("Inventory written → %s  (%d tables, %d enabled, %d disabled)",
             path, len(tables),
             sum(1 for t in tables if t.enabled),
             sum(1 for t in tables if not t.enabled))


def toggle_table(
    tables: List[TableInfo],
    fqn: str,
    *,
    enable: bool,
    reason: str = "",
) -> TableInfo:
    """
    Enable or disable a single table by FQN (SCHEMA.TABLE).

    Returns the modified TableInfo.
    Raises KeyError if table not found.
    """
    fqn_upper = fqn.strip().upper()
    for t in tables:
        if t.fqn == fqn_upper:
            t.enabled = enable
            if enable:
                t.disabled_reason = ""
                t.disabled_at = ""
                log.info("ENABLED  %s%s", t.fqn, f" (was: {reason})" if reason else "")
            else:
                t.disabled_reason = reason or "manually disabled"
                t.disabled_at = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
                log.info("DISABLED %s — %s", t.fqn, t.disabled_reason)
            return t
    raise KeyError(f"Table not found in inventory: {fqn_upper}")
